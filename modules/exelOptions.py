from openpyxl import (
    load_workbook
)
from PyQt5.QtWidgets import (
    QTableWidgetItem
)
from PyQt5.QtGui import (
    QColor
)
from PyQt5.QtCore import (
    Qt
)
from .dialog_message import Dialog

def loadExel(table, file):
    exel = load_workbook(file)
    sheet = exel.active
    
    max_rows = sheet.max_row
    max_columns = sheet.max_column
    
    table.clear()
    
    table.setRowCount(max_rows)
    table.setColumnCount(max_columns)
    
    for row in range(1, max_rows + 1):
        for column in range(1, max_columns + 1):
            cell = sheet.cell(row=row, column=column).value
            item = QTableWidgetItem(str(cell))
            if row == 1 or column == 1:
                item.setBackground(QColor(50,50,50))
            item.setTextAlignment(Qt.AlignCenter)
            table.setItem(row - 1, column - 1, item)
    
    table.horizontalHeader().setVisible(False)
    table.verticalHeader().setVisible(False)
    
def saveExel(table, file_path, full_path):
    try:
        exel = load_workbook(file_path)
        sheet = exel.active
    except FileNotFoundError:
        print("Error leyendo el archivo")
    
    for row in range(table.rowCount()):
        for column in range(table.columnCount()):
            item = table.item(row, column)
            if item is not None:
                sheet.cell(row=row + 1, column=column + 1).value = item.text()
                    
    try:
        exel.save(file_path)
        dialog = Dialog("Guardado Exitosamente", full_path)
        dialog.exec_()
    except:
        print("ERROR")